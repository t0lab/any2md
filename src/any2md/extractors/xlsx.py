"""xlsx extractor — openpyxl based, pure-Python.

Features:
- Multi-table detection via blank-row clustering
- Merged title rows surfaced as standalone_cells (not absorbed into tables)
- Header detection picks first non-merged-across-width row with distinct cells
"""
from __future__ import annotations

import logging
import re
import zipfile
from pathlib import Path
from typing import Any

from lxml import etree
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from any2md.extractors._image_utils import image_dimensions, normalize_to_png
from any2md.extractors._smartart import (
    detect_layout_kind,
    parse_rels,
    parse_smartart_tree,
)
from any2md.extractors.base import ExtractResult

log = logging.getLogger(__name__)

_NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "dgm": "http://schemas.openxmlformats.org/drawingml/2006/diagram",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}


class XlsxExtractor:
    format = "xlsx"

    def extract(self, source: Path) -> ExtractResult:
        warnings: list[str] = []
        images: dict[str, bytes] = {}
        wb = load_workbook(source, data_only=True)

        # Pre-load diagram + drawing artifacts for SmartArt + other drawing shapes
        self._dgm_data_xmls: dict[str, bytes] = {}
        self._dgm_layout_xmls: dict[str, bytes] = {}
        self._sheet_smartarts: dict[int, list[dict[str, Any]]] = {}
        # sheet_idx (1-based) → list[{name, anchor, geom, text, shape_kind}]
        self._sheet_drawing_shapes: dict[int, list[dict[str, Any]]] = {}
        try:
            self._preload_drawings(source, warnings)
        except Exception as e:
            warnings.append(f"Drawing preload failed: {e}")

        sheets_ir = []
        img_counter = 0
        for sheet_idx, ws in enumerate(wb.worksheets):
            if ws.sheet_state != "visible":
                warnings.append(f"Sheet {ws.title!r}: not visible ({ws.sheet_state}), skipped")
                continue

            # Build merged-cell value map (top-left value → all covered coords)
            merged_value_map: dict[str, Any] = {}
            merged_full_width: set[int] = set()  # row indices that are merged across full width
            full_width_cols = ws.max_column or 0
            for rng in ws.merged_cells.ranges:
                tl = ws.cell(row=rng.min_row, column=rng.min_col).value
                for r in range(rng.min_row, rng.max_row + 1):
                    for c in range(rng.min_col, rng.max_col + 1):
                        merged_value_map[f"{get_column_letter(c)}{r}"] = tl
                # Mark "title row" candidates: merged across the full used width on a single row
                if (rng.min_row == rng.max_row and
                        rng.min_col == 1 and rng.max_col == full_width_cols):
                    merged_full_width.add(rng.min_row)

            # Gather all raw rows with metadata
            raw_rows = self._gather_rows(ws, merged_value_map)

            # Cluster rows (blank-row separated)
            clusters = self._cluster_rows(raw_rows)

            # Per cluster: classify and emit tables vs standalone_cells
            tables: list[dict[str, Any]] = []
            standalone_cells: list[dict[str, Any]] = []
            for cluster in clusters:
                table_ir, label_cells = self._cluster_to_table(
                    cluster, merged_full_width, full_width_cols
                )
                if table_ir is not None:
                    tables.append(table_ir)
                standalone_cells.extend(label_cells)

            # Charts
            charts_ir = []
            for chart in getattr(ws, "_charts", []):
                charts_ir.append(self._extract_chart(chart))

            # Images
            images_anchors = []
            for img in getattr(ws, "_images", []):
                img_counter += 1
                blob = self._image_blob(img)
                if blob is None:
                    warnings.append(f"Sheet {ws.title!r}: image #{img_counter} could not be read")
                    continue
                normalized = normalize_to_png(blob)
                w_px, h_px = image_dimensions(normalized)
                image_id = f"img_s{sheet_idx + 1}_{img_counter}"
                rel_path = f"raw_images/{image_id}.png"
                images[rel_path] = normalized
                anchor_info = self._image_anchor_info(img)
                anchor_entry: dict[str, Any] = {
                    "image": {
                        "id": image_id,
                        "relative_path": rel_path,
                        "width_px": w_px,
                        "height_px": h_px,
                        "container": {
                            "kind": "xlsx_sheet",
                            "index": sheet_idx + 1,
                            "label": ws.title,
                        },
                        "neighbors": {
                            "before_text": "",
                            "after_text": "",
                            "before_kind": None,
                            "after_kind": None,
                        },
                    },
                    "anchor_cell": anchor_info.get("from_cell") or "A1",
                }
                if anchor_info.get("anchor_range"):
                    anchor_entry["anchor_range"] = anchor_info["anchor_range"]
                if anchor_info.get("anchor_kind"):
                    anchor_entry["anchor_kind"] = anchor_info["anchor_kind"]
                if anchor_info.get("extent_px"):
                    anchor_entry["extent_px"] = anchor_info["extent_px"]
                images_anchors.append(anchor_entry)

            # SmartArt shapes attached to this sheet
            smartart_ir = self._build_sheet_smartart_ir(
                sheet_idx + 1, ws.title, warnings
            )

            # Cell hyperlinks + comments (collected separately to avoid bloating tables)
            cell_hyperlinks, cell_comments = self._collect_cell_extras(ws)

            sheet_ir: dict[str, Any] = {
                "name": ws.title,
                "index": sheet_idx + 1,
                "used_range": ws.dimensions or "A1:A1",
                "tables": tables,
                "standalone_cells": standalone_cells,
                "charts": charts_ir,
                "images": images_anchors,
                "smartart": smartart_ir,
            }
            if cell_hyperlinks:
                sheet_ir["cell_hyperlinks"] = cell_hyperlinks
            if cell_comments:
                sheet_ir["cell_comments"] = cell_comments
            shapes_on_sheet = self._sheet_drawing_shapes.get(sheet_idx + 1, [])
            if shapes_on_sheet:
                sheet_ir["drawing_shapes"] = shapes_on_sheet
            sheets_ir.append(sheet_ir)

        ir = {
            "format": "xlsx",
            "source": str(source),
            "sheets": sheets_ir,
        }
        return ExtractResult(ir=ir, images=images, warnings=warnings)

    # ----------------- row gathering & clustering -----------------

    def _gather_rows(
        self, ws: Worksheet, merged_value_map: dict[str, Any]
    ) -> list[dict[str, Any]]:
        """Return rows as list of dicts with row_idx, values, bold_count, distinct_count, merged_full_width."""
        rows_out: list[dict[str, Any]] = []
        if not ws.max_row or not ws.max_column:
            return rows_out

        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row,
            min_col=1, max_col=ws.max_column,
            values_only=False,
        ):
            row_idx = row[0].row if row else 0
            values: list[Any] = []
            bold_count = 0
            for cell in row:
                if isinstance(cell, MergedCell):
                    values.append(merged_value_map.get(cell.coordinate))
                else:
                    values.append(cell.value)
                    if cell.font and cell.font.bold:
                        bold_count += 1

            non_null = [v for v in values if v is not None and v != ""]
            distinct_count = len(set(str(v) for v in non_null))
            is_blank = len(non_null) == 0
            is_merged_title = (
                len(non_null) > 0
                and len(set(str(v) for v in values if v is not None)) == 1
                and all(v is not None for v in values)
            )

            rows_out.append({
                "row_idx": row_idx,
                "values": values,
                "bold_count": bold_count,
                "non_null_count": len(non_null),
                "distinct_count": distinct_count,
                "is_blank": is_blank,
                "is_merged_title": is_merged_title,
            })
        return rows_out

    def _cluster_rows(self, raw_rows: list[dict[str, Any]]) -> list[list[dict[str, Any]]]:
        clusters: list[list[dict[str, Any]]] = []
        current: list[dict[str, Any]] = []
        for r in raw_rows:
            if r["is_blank"]:
                if current:
                    clusters.append(current)
                    current = []
            else:
                current.append(r)
        if current:
            clusters.append(current)
        return clusters

    def _cluster_to_table(
        self,
        cluster: list[dict[str, Any]],
        merged_full_width_rows: set[int],
        full_width_cols: int,
    ) -> tuple[dict[str, Any] | None, list[dict[str, Any]]]:
        """Convert a row cluster into either:
        - A table (with header detection + leading title rows extracted into standalone_cells)
        - Or pure standalone_cells (if no data rows present)
        """
        if not cluster:
            return None, []

        standalone: list[dict[str, Any]] = []

        # Leading merged-title rows → standalone label cells, NOT part of the table
        leading_strip = 0
        for r in cluster:
            if r["row_idx"] in merged_full_width_rows or r["is_merged_title"]:
                first_val = next(
                    (v for v in r["values"] if v is not None and v != ""), None
                )
                if first_val is not None:
                    standalone.append({
                        "coordinate": f"A{r['row_idx']}",
                        "value": first_val,
                        "kind": "title",
                    })
                leading_strip += 1
            else:
                break

        data_rows = cluster[leading_strip:]
        if not data_rows:
            # Cluster is entirely title-ish → standalone only
            return None, standalone

        # If there's exactly one data row → standalone label
        if len(data_rows) == 1:
            r = data_rows[0]
            first_val = next((v for v in r["values"] if v is not None and v != ""), None)
            if first_val is not None:
                standalone.append({
                    "coordinate": f"A{r['row_idx']}",
                    "value": first_val,
                    "kind": "note",
                })
            return None, standalone

        # Detect header row within data rows
        header_row_offset = self._pick_header_row(data_rows, merged_full_width_rows)
        header_row: list[str] | None = None
        if header_row_offset is not None:
            header_row = [
                str(v) if v is not None else ""
                for v in data_rows[header_row_offset]["values"]
            ]
            # Trim trailing empty header cells
            while header_row and header_row[-1] == "":
                header_row.pop()
            if not header_row:
                header_row = None

        rows_data = [r["values"] for r in data_rows]

        first_row_idx = data_rows[0]["row_idx"]
        last_row_idx = data_rows[-1]["row_idx"]
        cols_letter = get_column_letter(max(1, full_width_cols))
        table_range = f"A{first_row_idx}:{cols_letter}{last_row_idx}"

        return (
            {
                "range": table_range,
                "header_row": header_row,
                "rows": rows_data,
            },
            standalone,
        )

    def _pick_header_row(
        self,
        data_rows: list[dict[str, Any]],
        merged_full_width_rows: set[int],
    ) -> int | None:
        """Pick the offset (within data_rows) of the best header candidate.

        Preference order:
        1. First non-merged-title row that has ALL distinct non-null values
        2. First row with bold styling on ≥3 cells
        3. First non-blank row
        """
        for i, r in enumerate(data_rows):
            if r["row_idx"] in merged_full_width_rows or r["is_merged_title"]:
                continue
            if r["distinct_count"] >= 3 and r["distinct_count"] == r["non_null_count"]:
                return i

        for i, r in enumerate(data_rows):
            if r["bold_count"] >= 3:
                return i

        return 0 if data_rows else None

    # ----------------- charts & images -----------------

    def _extract_chart(self, chart: Any) -> dict[str, Any]:
        chart_type = self._guess_chart_type(chart)
        title = None
        try:
            if chart.title and chart.title.tx and chart.title.tx.rich:
                title = "".join(
                    r.t for p in chart.title.tx.rich.paragraphs for r in p.r if r.t
                )
        except Exception:
            pass

        categories: list[str] = []
        series_ir: list[dict[str, Any]] = []
        try:
            for ser in getattr(chart, "series", []) or []:
                name = None
                try:
                    if ser.tx and ser.tx.strRef and ser.tx.strRef.f:
                        name = ser.tx.strRef.f
                except Exception:
                    pass
                values: list[Any] = []
                try:
                    if ser.val and ser.val.numRef and ser.val.numRef.numCache:
                        values = [p.v for p in ser.val.numRef.numCache.pt]
                except Exception:
                    pass
                if not categories:
                    try:
                        if ser.cat and ser.cat.strRef and ser.cat.strRef.strCache:
                            categories = [p.v for p in ser.cat.strRef.strCache.pt]
                        elif ser.cat and ser.cat.numRef and ser.cat.numRef.numCache:
                            categories = [str(p.v) for p in ser.cat.numRef.numCache.pt]
                    except Exception:
                        pass
                series_ir.append({"name": name, "values": values})
        except Exception:
            pass

        return {
            "order": 0,
            "type": "chart",
            "bbox": {"left": 0, "top": 0, "width": 0, "height": 0, "unit": "emu"},
            "chart_type": chart_type,
            "title": title,
            "axis_labels": {},
            "categories": categories,
            "series": series_ir,
        }

    def _guess_chart_type(self, chart: Any) -> str:
        cls_name = type(chart).__name__.lower()
        for key in ("bar", "column", "line", "pie", "scatter", "area", "doughnut"):
            if key in cls_name:
                return key
        return "other"

    def _image_blob(self, img: Any) -> bytes | None:
        data = getattr(img, "_data", None)
        if callable(data):
            try:
                return data()
            except Exception:
                return None
        if isinstance(data, (bytes, bytearray)):
            return bytes(data)
        ref = getattr(img, "ref", None) or getattr(img, "path", None)
        if ref and hasattr(ref, "read"):
            try:
                return ref.read()
            except Exception:
                return None
        return None

    # ----------------- Drawings (SmartArt + autoshapes + connectors) -----------------

    def _preload_drawings(self, source: Path, warnings: list[str]) -> None:
        """Scan the xlsx zip for drawings attached to each sheet.

        Handles three drawing child types:
          - xdr:graphicFrame with dgm:relIds → SmartArt
          - xdr:sp                            → text boxes / autoshapes
          - xdr:cxnSp                         → connectors / arrows
        """
        with zipfile.ZipFile(source) as z:
            zip_files = {name: z.read(name) for name in z.namelist()
                         if name.startswith("xl/diagrams/") or
                            name.startswith("xl/drawings/") or
                            name.startswith("xl/worksheets/_rels/")}

        for name, raw in zip_files.items():
            if name.startswith("xl/diagrams/data") and name.endswith(".xml"):
                self._dgm_data_xmls[name] = raw
            elif name.startswith("xl/diagrams/layout") and name.endswith(".xml"):
                self._dgm_layout_xmls[name] = raw

        # Map sheet → ALL drawing targets (a sheet can reference multiple drawings)
        sheet_to_drawings: dict[int, list[str]] = {}
        for name, raw in zip_files.items():
            m = re.match(r"xl/worksheets/_rels/sheet(\d+)\.xml\.rels$", name)
            if not m:
                continue
            sheet_idx = int(m.group(1))
            rels_map = parse_rels(raw)
            for _, target in rels_map.items():
                if "drawings/" not in target or not target.endswith(".xml"):
                    continue
                if target.startswith("/"):
                    norm = target.lstrip("/")
                elif target.startswith("../"):
                    norm = "xl/" + target[3:]
                else:
                    norm = target
                sheet_to_drawings.setdefault(sheet_idx, []).append(norm)

        for sheet_idx, drawing_paths in sheet_to_drawings.items():
            for drawing_path in drawing_paths:
                drawing_xml = zip_files.get(drawing_path)
                if drawing_xml is None:
                    continue
                drawing_rels_path = (
                    drawing_path.rsplit("/", 1)[0] + "/_rels/" +
                    drawing_path.rsplit("/", 1)[1] + ".rels"
                )
                drawing_rels_xml = zip_files.get(drawing_rels_path, b"")
                drawing_rels = parse_rels(drawing_rels_xml) if drawing_rels_xml else {}

                try:
                    tree = etree.fromstring(drawing_xml)
                except Exception as e:
                    warnings.append(f"Drawing {drawing_path!r} parse failed: {e}")
                    continue

                self._parse_drawing_tree(
                    tree, sheet_idx, drawing_path, drawing_rels, warnings
                )

    def _parse_drawing_tree(
        self,
        tree: Any,
        sheet_idx: int,
        drawing_path: str,
        drawing_rels: dict[str, str],
        warnings: list[str],
    ) -> None:
        # SmartArt (graphicFrame with dgm:relIds)
        for gf in tree.xpath("//xdr:graphicFrame", namespaces=_NS):
            gd = gf.find(".//a:graphicData", namespaces=_NS)
            if gd is None or "diagram" not in (gd.get("uri") or ""):
                continue
            rel = gd.find(".//dgm:relIds", namespaces=_NS)
            if rel is None:
                continue
            dm_rid = rel.get(f"{{{_NS['r']}}}dm")
            lo_rid = rel.get(f"{{{_NS['r']}}}lo")
            data_xml = None
            layout_xml = None
            if dm_rid and dm_rid in drawing_rels:
                data_target = self._normalize_drawing_rel_target(
                    drawing_rels[dm_rid], drawing_path
                )
                data_xml = self._dgm_data_xmls.get(data_target)
            if lo_rid and lo_rid in drawing_rels:
                layout_target = self._normalize_drawing_rel_target(
                    drawing_rels[lo_rid], drawing_path
                )
                layout_xml = self._dgm_layout_xmls.get(layout_target)

            anchor_info = self._extract_drawing_anchor(gf)
            name_attrs = gf.xpath(".//xdr:cNvPr/@name", namespaces=_NS)
            shape_name = name_attrs[0] if name_attrs else "SmartArt"
            self._sheet_smartarts.setdefault(sheet_idx, []).append({
                "data_xml": data_xml,
                "layout_xml": layout_xml,
                "anchor": anchor_info,
                "name": shape_name,
            })

        # Autoshapes / text boxes (xdr:sp)
        for sp in tree.xpath("//xdr:sp", namespaces=_NS):
            name_attrs = sp.xpath(".//xdr:cNvPr/@name", namespaces=_NS)
            name = name_attrs[0] if name_attrs else "Shape"
            geom_attrs = sp.xpath(".//a:prstGeom/@prst", namespaces=_NS)
            geom = geom_attrs[0] if geom_attrs else None
            text = self._extract_shape_text(sp)
            anchor_info = self._extract_drawing_anchor(sp)
            anchor_range = self._anchor_to_range(anchor_info)
            # Classify: pure text box has no preset geom or geom == "rect" with txBody
            shape_kind = "textbox"
            if geom and geom not in ("rect", "roundRect", "ellipse"):
                if "arrow" in geom.lower() or "callout" in geom.lower():
                    shape_kind = "arrow"
                else:
                    shape_kind = "other"
            if not text and shape_kind == "textbox":
                # skip empty decorative shapes
                continue
            self._sheet_drawing_shapes.setdefault(sheet_idx, []).append({
                "name": name,
                "anchor_range": anchor_range,
                "geom": geom,
                "text": text or None,
                "shape_kind": shape_kind,
            })

        # Connectors (xdr:cxnSp) — lines/arrows linking shapes
        for cxn in tree.xpath("//xdr:cxnSp", namespaces=_NS):
            name_attrs = cxn.xpath(".//xdr:cNvPr/@name", namespaces=_NS)
            name = name_attrs[0] if name_attrs else "Connector"
            geom_attrs = cxn.xpath(".//a:prstGeom/@prst", namespaces=_NS)
            geom = geom_attrs[0] if geom_attrs else None
            anchor_info = self._extract_drawing_anchor(cxn)
            anchor_range = self._anchor_to_range(anchor_info)
            self._sheet_drawing_shapes.setdefault(sheet_idx, []).append({
                "name": name,
                "anchor_range": anchor_range,
                "geom": geom,
                "text": None,
                "shape_kind": "connector",
            })

    def _extract_shape_text(self, sp_element: Any) -> str:
        """Extract concatenated text from an xdr:sp <a:txBody>."""
        parts: list[str] = []
        for p in sp_element.xpath(".//a:txBody/a:p", namespaces=_NS):
            run_texts = [t.text or "" for t in p.xpath(".//a:t", namespaces=_NS)]
            line = "".join(run_texts).strip()
            if line:
                parts.append(line)
        return "\n".join(parts)

    def _anchor_to_range(self, anchor: dict[str, Any] | None) -> str | None:
        if not anchor:
            return None
        fc = get_column_letter(anchor["from_col"])
        tc = get_column_letter(anchor["to_col"])
        return f"{fc}{anchor['from_row']}:{tc}{anchor['to_row']}"

    def _normalize_drawing_rel_target(self, target: str, drawing_path: str) -> str:
        if target.startswith("/"):
            return target.lstrip("/")
        if target.startswith("../"):
            # Walk up from drawing_path's directory
            parts = drawing_path.rsplit("/", 1)[0].split("/")
            rest = target
            while rest.startswith("../"):
                if parts:
                    parts.pop()
                rest = rest[3:]
            return "/".join(parts + [rest]) if parts else rest
        return target

    def _extract_drawing_anchor(self, gf_element: Any) -> dict[str, Any] | None:
        """Walk up from the graphicFrame to find its enclosing anchor (twoCellAnchor or oneCellAnchor).
        Returns {from_col, from_row, to_col, to_row} (1-based) or None.
        """
        parent = gf_element.getparent()
        while parent is not None:
            tag = parent.tag.split("}")[-1]
            if tag in ("twoCellAnchor", "oneCellAnchor"):
                from_el = parent.find("xdr:from", namespaces=_NS)
                to_el = parent.find("xdr:to", namespaces=_NS)
                if from_el is None:
                    return None
                fc = int(from_el.findtext("xdr:col", default="0", namespaces=_NS))
                fr = int(from_el.findtext("xdr:row", default="0", namespaces=_NS))
                tc = int(to_el.findtext("xdr:col", default="0", namespaces=_NS)) if to_el is not None else fc
                tr = int(to_el.findtext("xdr:row", default="0", namespaces=_NS)) if to_el is not None else fr
                return {
                    "from_col": fc + 1, "from_row": fr + 1,
                    "to_col": tc + 1, "to_row": tr + 1,
                }
            parent = parent.getparent()
        return None

    def _build_sheet_smartart_ir(
        self, sheet_idx: int, sheet_name: str, warnings: list[str]
    ) -> list[dict[str, Any]]:
        ir_list: list[dict[str, Any]] = []
        for sa in self._sheet_smartarts.get(sheet_idx, []):
            nodes: list[dict[str, Any]] = []
            if sa["data_xml"] is not None:
                try:
                    nodes = parse_smartart_tree(sa["data_xml"])
                except Exception as e:
                    warnings.append(
                        f"Sheet {sheet_name!r}: SmartArt {sa['name']!r} data parse failed: {e}"
                    )
            layout_kind = "other"
            if sa["layout_xml"] is not None:
                layout_kind = detect_layout_kind(sa["layout_xml"])

            anchor = sa.get("anchor")
            anchor_range = None
            if anchor:
                fc = get_column_letter(anchor["from_col"])
                tc = get_column_letter(anchor["to_col"])
                anchor_range = f"{fc}{anchor['from_row']}:{tc}{anchor['to_row']}"

            ir_list.append({
                "type": "smartart",
                "name": sa["name"],
                "anchor_range": anchor_range,
                "layout_kind": layout_kind,
                "nodes": nodes,
            })
        return ir_list

    # ----------------- cell extras -----------------

    def _collect_cell_extras(
        self, ws: Worksheet
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        """Walk the worksheet cells once and collect hyperlinks + comments."""
        hyperlinks: list[dict[str, Any]] = []
        comments: list[dict[str, Any]] = []
        if not ws.max_row or not ws.max_column:
            return hyperlinks, comments
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                 min_col=1, max_col=ws.max_column,
                                 values_only=False):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                link = getattr(cell, "hyperlink", None)
                if link is not None:
                    target = (
                        getattr(link, "target", None)
                        or getattr(link, "location", None)
                        or getattr(link, "ref", None)
                    )
                    if target:
                        entry: dict[str, Any] = {
                            "coordinate": cell.coordinate,
                            "target": str(target),
                        }
                        if cell.value not in (None, ""):
                            entry["display"] = str(cell.value)
                        hyperlinks.append(entry)
                comment = getattr(cell, "comment", None)
                if comment is not None and comment.text:
                    comments.append({
                        "coordinate": cell.coordinate,
                        "author": comment.author,
                        "text": comment.text,
                    })
        return hyperlinks, comments

    # ----------------- helpers -----------------

    def _image_anchor_info(self, img: Any) -> dict[str, Any]:
        """Resolve image anchor → {from_cell, anchor_range, anchor_kind, extent_emu}.

        openpyxl anchor types:
          - OneCellAnchor: _from + ext (cx, cy in EMU)
          - TwoCellAnchor: _from + to (both AnchorMarker)
          - AbsoluteAnchor: pos + ext (no cells)
        """
        out: dict[str, Any] = {}
        try:
            anc = img.anchor
        except Exception:
            return out
        if anc is None:
            return out

        cls_name = type(anc).__name__.lower()
        if "twocell" in cls_name:
            kind = "two_cell"
        elif "onecell" in cls_name:
            kind = "one_cell"
        elif "absolute" in cls_name:
            kind = "absolute"
        else:
            kind = "one_cell"
        out["anchor_kind"] = kind

        from_attr = getattr(anc, "_from", None) or getattr(anc, "from", None)
        if from_attr is not None:
            fc = getattr(from_attr, "col", 0) + 1
            fr = getattr(from_attr, "row", 0) + 1
            out["from_cell"] = f"{get_column_letter(fc)}{fr}"

            to_attr = getattr(anc, "to", None)
            if to_attr is not None:
                tc = getattr(to_attr, "col", fc - 1) + 1
                tr = getattr(to_attr, "row", fr - 1) + 1
                out["anchor_range"] = (
                    f"{get_column_letter(fc)}{fr}:{get_column_letter(tc)}{tr}"
                )

        ext = getattr(anc, "ext", None)
        if ext is not None:
            cx = getattr(ext, "cx", None)
            cy = getattr(ext, "cy", None)
            if cx is not None and cy is not None:
                out["extent_px"] = {
                    "cx": round(int(cx) / 9525, 1),
                    "cy": round(int(cy) / 9525, 1),
                }

        return out
